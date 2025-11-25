from pathlib import Path
from io import BytesIO

from openpyxl import load_workbook
from PIL import Image


# ใช้ไฟล์ตัวเดียวกับที่แอปอ่านอยู่
EXCEL_FILE = "Spare parts list for TOA- Combine.xlsx"


def main():
    base = Path(__file__).parent
    xlsx_path = base / EXCEL_FILE

    if not xlsx_path.exists():
        print(f"ไม่พบไฟล์: {xlsx_path}")
        return

    print(f"โหลดไฟล์: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)

    # โฟลเดอร์สำหรับเก็บรูป
    images_dir = base / "images"
    spare_dir = images_dir / "spare"
    product_dir = images_dir / "product"
    spare_dir.mkdir(parents=True, exist_ok=True)
    product_dir.mkdir(parents=True, exist_ok=True)

    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title} ---")

        HEADER_ROW = 2  # แถวที่มีหัวคอลัมน์จริง (Warranty, Model, Product Name, ...)

        # map: column_index -> header text
        headers = {}
        for cell in ws[HEADER_ROW]:
            if cell.value is not None:
                headers[cell.column] = str(cell.value)

        def find_col(keyword: str):
            keyword = keyword.lower()
            for col_idx, title in headers.items():
                if keyword in str(title).lower():
                    return col_idx
            return None

        col_code = find_col("spare part code")
        col_model = find_col("model")
        col_pname = find_col("product name")

        print(
            "คอลัมน์ที่ใช้:",
            "Spare Part Code =", col_code,
            "| Model =", col_model,
            "| Product Name =", col_pname,
        )

        # ฟังก์ชันไล่ย้อนขึ้นไปหา Model / Product Name ที่ใกล้ที่สุดด้านบน
        def get_model_pname(start_row: int):
            for r in range(start_row, HEADER_ROW, -1):
                model = ws.cell(row=r, column=col_model).value if col_model else None
                pname = ws.cell(row=r, column=col_pname).value if col_pname else None
                if (model and str(model).strip()) or (pname and str(pname).strip()):
                    return model, pname
            return None, None

        prod_count = 0
        spare_count = 0

        # loop รูปที่ฝังทั้งหมดในชีต
        for img in getattr(ws, "_images", []):
            anchor = img.anchor._from  # 0-based index
            row = anchor.row + 1       # แปลงเป็น row แบบ Excel (เริ่มที่ 1)
            col = anchor.col + 1

            header_text = str(ws.cell(row=HEADER_ROW, column=col).value or "")
            header_lower = header_text.lower()

            # แปลงรูปจาก openpyxl เป็น PIL image
            blob = img._data()
            if callable(blob):
                blob = blob()  # บางเวอร์ชัน _data() คืนฟังก์ชัน
            try:
                pil_img = Image.open(BytesIO(blob))
            except Exception as e:
                print(f"  ข้ามรูป (เปิดไม่ได้): row {row}, col {col}, error = {e}")
                continue

            # ---------- รูป Spare part ----------
            if "spare" in header_lower:
                if not col_code:
                    continue
                spare_code = ws.cell(row=row, column=col_code).value
                if not spare_code:
                    continue

                filename = f"{str(spare_code).strip().replace('/', '_')}.png"
                target = spare_dir / filename
                pil_img.save(target)
                spare_count += 1
                # print(f"  [Spare]  row {row} -> {target.name}")

            # ---------- รูป Product ----------
            elif "product" in header_lower or ("picture" in header_lower and "spare" not in header_lower):
                model, pname = get_model_pname(row)

                base_name = None
                if model and str(model).strip():
                    base_name = str(model).strip()
                elif pname and str(pname).strip():
                    base_name = str(pname).strip()
                else:
                    # หา model/pname ไม่เจอ ข้าม
                    continue

                safe = "".join(ch if ch.isalnum() else "_" for ch in base_name)
                filename = f"{safe}.png"
                target = product_dir / filename
                pil_img.save(target)
                prod_count += 1
                # print(f"  [Product] row {row} -> {target.name}")

            # ไม่ใช่ Product / Spare ข้าม
            else:
                continue

        print(f"เซฟรูป Product {prod_count} รูป | Spare {spare_count} รูป")

    print("\nเสร็จแล้ว! รูปถูกเซฟไว้ที่โฟลเดอร์:")
    print(f"  - {spare_dir}")
    print(f"  - {product_dir}")


if __name__ == "__main__":
    main()
